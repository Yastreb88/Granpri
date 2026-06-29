from flask import Flask, render_template, request, redirect, url_for, session, send_file, flash, jsonify
import pandas as pd
import os
import sqlite3
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import hashlib
import config

app = Flask(__name__)
app.secret_key = 'supersecretkey'

UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# --- Инициализация базы данных ---
def init_db():
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS users
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  username TEXT UNIQUE,
                  password_hash TEXT,
                  password_plain TEXT,
                  role TEXT DEFAULT 'club_admin',
                  club TEXT)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS uploads
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  user_id INTEGER,
                  season TEXT,
                  club TEXT,
                  tour TEXT,
                  category TEXT,
                  filename TEXT,
                  uploaded_at DATETIME DEFAULT CURRENT_TIMESTAMP)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS clubs
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  name TEXT,
                  season TEXT,
                  UNIQUE(name, season))''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS tours
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  number TEXT,
                  season TEXT,
                  UNIQUE(number, season))''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS categories
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  name TEXT,
                  season TEXT,
                  UNIQUE(name, season))''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS seasons
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  name TEXT UNIQUE)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS settings
                 (key TEXT PRIMARY KEY,
                  value TEXT)''')
    
    c.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('top_tours', ?)", (str(config.TOP_TOURS),))
    
    admin = c.execute("SELECT * FROM users WHERE username='Admin'").fetchone()
    if not admin:
        password_hash = hashlib.sha256('admin123'.encode()).hexdigest()
        c.execute("INSERT INTO users (username, password_hash, password_plain, role, club) VALUES (?, ?, ?, ?, ?)",
                  ('Admin', password_hash, 'admin123', 'super_admin', ''))
    
    for season in config.SEASONS:
        c.execute("INSERT OR IGNORE INTO seasons (name) VALUES (?)", (season,))
    
    conn.commit()
    conn.close()

def load_defaults():
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    season = config.ACTIVE_SEASON
    
    for cat in config.CATEGORIES:
        c.execute("INSERT OR IGNORE INTO categories (name, season) VALUES (?, ?)", (cat, season))
    
    for club in config.CLUBS:
        c.execute("INSERT OR IGNORE INTO clubs (name, season) VALUES (?, ?)", (club, season))
    
    for tour in config.TOURS:
        c.execute("INSERT OR IGNORE INTO tours (number, season) VALUES (?, ?)", (tour, season))
    
    conn.commit()
    conn.close()

def get_top_tours():
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    result = c.execute("SELECT value FROM settings WHERE key='top_tours'").fetchone()
    conn.close()
    if result:
        try:
            return int(result[0])
        except:
            return config.TOP_TOURS
    return config.TOP_TOURS

def set_top_tours(value):
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute("UPDATE settings SET value=? WHERE key='top_tours'", (str(value),))
    conn.commit()
    conn.close()

init_db()
load_defaults()

# --- Таблица очков ---
POINTS = {
    1:60,2:54,3:48,4:43,5:40,6:38,7:36,8:34,9:32,10:31,
    11:30,12:29,13:28,14:27,15:26,16:25,17:24,18:23,19:22,20:21,
    21:20,22:19,23:18,24:17,25:16,26:15,27:14,28:13,29:12,30:11,
    31:10,32:9,33:8,34:7,35:6,36:5,37:4,38:3,39:2,40:1
}

def process_csv(file_path, club, tour, category):
    df = pd.read_csv(file_path, sep=';', encoding='utf-8-sig', header=None)
    df = df.iloc[:, :3].copy()
    df.columns = ['Место', 'ID', 'ФИО']
    df['Место'] = pd.to_numeric(df['Место'], errors='coerce')
    df['Очки'] = df['Место'].map(POINTS).fillna(0).astype(int)
    df['Турнир'] = f"Тур {tour}"

    name_to_ids = {}
    for _, row in df.iterrows():
        if pd.notna(row['ID']) and str(row['ID']).strip():
            if row['ФИО'] not in name_to_ids:
                name_to_ids[row['ФИО']] = set()
            name_to_ids[row['ФИО']].add(str(row['ID']))
    conflict_names = [name for name, ids in name_to_ids.items() if len(ids) > 1]
    df['Конфликт_ID'] = df['ФИО'].apply(lambda x: x in conflict_names)

    df['Уникальный_игрок'] = df['ID'].astype(str) + '|' + df['ФИО']
    pivot = df.pivot_table(
        index=['Уникальный_игрок', 'ID', 'ФИО', 'Конфликт_ID'],
        columns='Турнир',
        values='Очки',
        fill_value=0
    )
    top_n = get_top_tours()
    pivot[f'Сумма лучших {top_n}'] = pivot.apply(lambda row: sum(sorted(row.values, reverse=True)[:top_n]), axis=1)
    pivot = pivot.sort_values(f'Сумма лучших {top_n}', ascending=False)
    result = pivot.reset_index()
    result = result.drop(columns=['Уникальный_игрок'])

    safe_club = "".join(c for c in club if c.isalnum() or c in (' ', '-', '_')).strip()
    output_filename = f"kubok_{safe_club}_{category}_{tour}.xlsx"
    output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
    result.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    conflict_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == 'Конфликт_ID':
            conflict_col = col
            break
    if conflict_col:
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=conflict_col).value == True:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = red_fill
        ws.delete_cols(conflict_col)
    wb.save(output_path)
    return output_path

# --- Утилиты ---
def get_seasons():
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    seasons = c.execute("SELECT name FROM seasons").fetchall()
    conn.close()
    return [s[0] for s in seasons] if seasons else config.SEASONS

def get_clubs(season=None):
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    if season:
        clubs = c.execute("SELECT name FROM clubs WHERE season=?", (season,)).fetchall()
    else:
        clubs = c.execute("SELECT name FROM clubs").fetchall()
    conn.close()
    return [c[0] for c in clubs]

def get_tours(season=None):
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    if season:
        tours = c.execute("SELECT number FROM tours WHERE season=?", (season,)).fetchall()
    else:
        tours = c.execute("SELECT number FROM tours").fetchall()
    conn.close()
    return [t[0] for t in tours]

def get_categories(season=None):
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    if season:
        cats = c.execute("SELECT name FROM categories WHERE season=?", (season,)).fetchall()
    else:
        cats = c.execute("SELECT name FROM categories").fetchall()
    conn.close()
    return [c[0] for c in cats]

def get_users():
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    users = c.execute("SELECT username, password_plain, role, club FROM users").fetchall()
    conn.close()
    return [{'username': u[0], 'password': u[1] or 'не задан', 'role': u[2], 'club': u[3]} for u in users]

# --- Маршруты ---
@app.route('/')
def index():
    if 'user' in session:
        return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        user = c.execute("SELECT * FROM users WHERE username=? AND password_hash=?", (username, password_hash)).fetchone()
        conn.close()
        if user:
            session['user'] = username
            session['role'] = user[3]
            session['club'] = user[4]
            return redirect(url_for('dashboard'))
        else:
            flash('Неверный логин или пароль')
            return redirect(url_for('login'))
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
def dashboard():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    if session.get('user') == 'Admin':
        session['role'] = 'super_admin'
    
    if session.get('role') != 'super_admin':
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        club_from_db = c.execute("SELECT club FROM users WHERE username=?", (session['user'],)).fetchone()
        conn.close()
        if club_from_db and club_from_db[0]:
            session['club'] = club_from_db[0]
    
    role = session.get('role')
    
    if role == 'super_admin':
        seasons = get_seasons()
        selected_season = request.args.get('season', config.ACTIVE_SEASON)
        if selected_season not in seasons:
            selected_season = config.ACTIVE_SEASON
        
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        uploads = c.execute("SELECT id, season, club, tour, category, filename, uploaded_at FROM uploads ORDER BY uploaded_at DESC").fetchall()
        conn.close()
        
        clubs = get_clubs(selected_season)
        tours = get_tours(selected_season)
        categories = get_categories(selected_season)
        top_tours = get_top_tours()
        
        club_stats = []
        for club in clubs:
            club_data = {'club': club, 'loaded': 0, 'total': len(tours)}
            for tour in tours:
                conn = sqlite3.connect('database.db')
                c = conn.cursor()
                loaded = c.execute("SELECT COUNT(*) FROM uploads WHERE club=? AND tour=? AND season=?", 
                                   (club, tour, selected_season)).fetchone()[0]
                conn.close()
                if loaded > 0:
                    club_data['loaded'] += 1
            club_stats.append(club_data)
        
        progress_table = []
        for club in clubs:
            for tour in tours:
                row = {'club': club, 'tour': tour}
                for cat in categories:
                    conn = sqlite3.connect('database.db')
                    c = conn.cursor()
                    loaded = c.execute("SELECT COUNT(*) FROM uploads WHERE club=? AND tour=? AND category=? AND season=?", 
                                       (club, tour, cat, selected_season)).fetchone()[0]
                    conn.close()
                    row[cat] = loaded > 0
                progress_table.append(row)
        
        users = get_users()
        
        return render_template('admin.html', 
                             seasons=seasons,
                             active_season=selected_season,
                             clubs=clubs,
                             tours=tours,
                             categories=categories,
                             top_tours=top_tours,
                             uploads=uploads,
                             club_stats=club_stats,
                             progress_table=progress_table,
                             total_tours=len(tours),
                             users=users)
    else:
        user_club = session.get('club', '')
        seasons = get_seasons()
        selected_season = request.args.get('season', config.ACTIVE_SEASON)
        if selected_season not in seasons:
            selected_season = config.ACTIVE_SEASON
        
        club_list = [user_club] if user_club else get_clubs(selected_season)
        return render_template('club.html', 
                             seasons=seasons,
                             active_season=selected_season,
                             clubs=club_list,
                             tours=get_tours(selected_season),
                             categories=get_categories(selected_season))

@app.route('/upload', methods=['POST'])
def upload():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    season = request.form.get('season')
    if not season:
        season = config.ACTIVE_SEASON
    
    club = session.get('club', '')
    
    if not club:
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        club_from_db = c.execute("SELECT club FROM users WHERE username=?", (session['user'],)).fetchone()
        conn.close()
        if club_from_db and club_from_db[0]:
            club = club_from_db[0]
            session['club'] = club
    
    if not club:
        flash('❌ Ошибка: клуб не привязан к вашему аккаунту. Обратитесь к суперадмину.')
        return redirect(url_for('dashboard'))
    
    tour = request.form['tour']
    category = request.form['category']
    file = request.files['file']
    
    if file and file.filename.endswith('.csv'):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        user = c.execute("SELECT id FROM users WHERE username=?", (session['user'],)).fetchone()
        c.execute("INSERT INTO uploads (user_id, season, club, tour, category, filename) VALUES (?, ?, ?, ?, ?, ?)",
                  (user[0], season, club, tour, category, filename))
        conn.commit()
        conn.close()
        
        try:
            result_path = process_csv(filepath, club, tour, category)
            session['download_file'] = os.path.basename(result_path)
            flash('✅ Файл успешно отправлен!')
            return redirect(url_for('upload_success'))
        except Exception as e:
            flash(f'❌ Ошибка обработки CSV: {str(e)}')
            return redirect(url_for('dashboard'))
    else:
        flash('❌ Пожалуйста, загрузите CSV-файл')
        return redirect(url_for('dashboard'))

@app.route('/upload_success')
def upload_success():
    return render_template('upload_success.html')

@app.route('/download/<filename>')
def download_file(filename):
    return send_file(os.path.join(app.config['UPLOAD_FOLDER'], filename), as_attachment=True)

@app.route('/download_csv/<int:upload_id>')
def download_csv(upload_id):
    if 'user' not in session or session.get('role') != 'super_admin':
        return redirect(url_for('login'))
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    upload = c.execute("SELECT filename FROM uploads WHERE id=?", (upload_id,)).fetchone()
    conn.close()
    if upload:
        filename = upload[0]
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(filepath):
            return send_file(filepath, as_attachment=True, download_name=filename)
    flash('❌ CSV-файл не найден')
    return redirect(url_for('dashboard'))

@app.route('/download_excel/<int:upload_id>')
def download_excel(upload_id):
    if 'user' not in session or session.get('role') != 'super_admin':
        return redirect(url_for('login'))
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    upload = c.execute("SELECT club, tour, category, filename FROM uploads WHERE id=?", (upload_id,)).fetchone()
    conn.close()
    if upload:
        club, tour, category, csv_filename = upload
        safe_club = "".join(c for c in club if c.isalnum() or c in (' ', '-', '_')).strip()
        excel_filename = f"kubok_{safe_club}_{category}_{tour}.xlsx"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
        if os.path.exists(filepath):
            return send_file(filepath, as_attachment=True, download_name=excel_filename)
    flash('❌ Excel-файл не найден')
    return redirect(url_for('dashboard'))

@app.route('/generate_summary', methods=['POST'])
def generate_summary():
    if 'user' not in session or session.get('role') != 'super_admin':
        return redirect(url_for('login'))
    
    season = request.form['season']
    club = request.form['club']
    category = request.form['category']
    top_n = get_top_tours()
    
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    uploads = c.execute("SELECT tour, filename FROM uploads WHERE season=? AND club=? AND category=? ORDER BY tour", 
                        (season, club, category)).fetchall()
    conn.close()
    
    if not uploads:
        flash(f'❌ Нет загруженных файлов для сезона "{season}", клуба "{club}" и категории "{category}"')
        return redirect(url_for('dashboard'))
    
    all_data = []
    for tour, filename in uploads:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(filepath):
            df = pd.read_csv(filepath, sep=';', encoding='utf-8-sig', header=None)
            df = df.iloc[:, :3].copy()
            df.columns = ['Место', 'ID', 'ФИО']
            df['Место'] = pd.to_numeric(df['Место'], errors='coerce')
            df['Очки'] = df['Место'].map(POINTS).fillna(0).astype(int)
            df['Турнир'] = f"Тур {tour}"
            all_data.append(df)
    
    if not all_data:
        flash('❌ Не удалось прочитать файлы')
        return redirect(url_for('dashboard'))
    
    combined = pd.concat(all_data, ignore_index=True)
    
    # Строим сводную таблицу
    pivot = combined.pivot_table(
        index=['ID', 'ФИО'],
        columns='Турнир',
        values='Очки',
        fill_value=0
    )
    pivot = pivot.reindex(sorted(pivot.columns), axis=1)
    
    # Выявление конфликтов
    name_to_ids = {}
    for _, row in combined.iterrows():
        if pd.notna(row['ID']) and str(row['ID']).strip():
            if row['ФИО'] not in name_to_ids:
                name_to_ids[row['ФИО']] = set()
            name_to_ids[row['ФИО']].add(str(row['ID']))
    conflict_names = [name for name, ids in name_to_ids.items() if len(ids) > 1]
    
    pivot['Конфликт_ID'] = pivot.index.get_level_values('ФИО').isin(conflict_names)
    
    # Сумма лучших N туров
    pivot[f'Сумма лучших {top_n}'] = pivot.apply(
        lambda row: sum(sorted(row.values[:-1], reverse=True)[:top_n]), 
        axis=1
    )
    
    pivot = pivot.sort_values(f'Сумма лучших {top_n}', ascending=False)
    result = pivot.reset_index()
    
    # Сохраняем Excel
    safe_club = "".join(c for c in club if c.isalnum() or c in (' ', '-', '_')).strip()
    output_filename = f"summary_{safe_club}_{category}.xlsx"
    output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
    result.to_excel(output_path, index=False)
    
    # --- Применяем стили ---
    wb = load_workbook(output_path)
    ws = wb.active
    
    sum_col_name = f'Сумма лучших {top_n}'
    sum_col_idx = None
    conflict_col_idx = None
    
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value == sum_col_name:
            sum_col_idx = col
        if cell_value == 'Конфликт_ID':
            conflict_col_idx = col
    
    # Жирный шрифт для суммы
    if sum_col_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=sum_col_idx)
            cell.font = Font(bold=True)
    
    # Красная заливка для конфликтов
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    if conflict_col_idx:
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=conflict_col_idx).value == True:
                for col in range(1, ws.max_column + 1):
                    if col != sum_col_idx:
                        ws.cell(row=row, column=col).fill = red_fill
    
    # Удаляем служебную колонку
    if conflict_col_idx:
        ws.delete_cols(conflict_col_idx)
    
    wb.save(output_path)
    
    flash('✅ Итоговая таблица успешно сформирована!')
    return send_file(output_path, as_attachment=True, download_name=os.path.basename(output_path))

# --- API ---
@app.route('/api/clubs', methods=['GET', 'POST', 'DELETE'])
def api_clubs():
    if 'user' not in session or session.get('role') != 'super_admin':
        return jsonify({'error': 'Unauthorized'}), 401
    if request.method == 'GET':
        season = request.args.get('season', config.ACTIVE_SEASON)
        return jsonify(get_clubs(season))
    elif request.method == 'POST':
        data = request.json
        name = data.get('name')
        season = data.get('season', config.ACTIVE_SEASON)
        if not name:
            return jsonify({'error': 'Название обязательно'}), 400
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        try:
            c.execute("INSERT OR IGNORE INTO clubs (name, season) VALUES (?, ?)", (name, season))
            conn.commit()
            conn.close()
            return jsonify({'status': 'ok'})
        except Exception as e:
            conn.close()
            return jsonify({'error': str(e)}), 400
    elif request.method == 'DELETE':
        data = request.json
        name = data.get('name')
        if not name:
            return jsonify({'error': 'Название обязательно'}), 400
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        c.execute("DELETE FROM clubs WHERE name=? AND season=?", (name, config.ACTIVE_SEASON))
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})

@app.route('/api/tours', methods=['GET', 'POST', 'DELETE'])
def api_tours():
    if 'user' not in session or session.get('role') != 'super_admin':
        return jsonify({'error': 'Unauthorized'}), 401
    if request.method == 'GET':
        season = request.args.get('season', config.ACTIVE_SEASON)
        return jsonify(get_tours(season))
    elif request.method == 'POST':
        data = request.json
        number = data.get('number')
        season = data.get('season', config.ACTIVE_SEASON)
        if not number:
            return jsonify({'error': 'Номер тура обязателен'}), 400
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        try:
            c.execute("INSERT OR IGNORE INTO tours (number, season) VALUES (?, ?)", (number, season))
            conn.commit()
            conn.close()
            return jsonify({'status': 'ok'})
        except Exception as e:
            conn.close()
            return jsonify({'error': str(e)}), 400
    elif request.method == 'DELETE':
        data = request.json
        number = data.get('number')
        if not number:
            return jsonify({'error': 'Номер тура обязателен'}), 400
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        c.execute("DELETE FROM tours WHERE number=? AND season=?", (number, config.ACTIVE_SEASON))
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})

@app.route('/api/categories', methods=['GET', 'POST', 'DELETE'])
def api_categories():
    if 'user' not in session or session.get('role') != 'super_admin':
        return jsonify({'error': 'Unauthorized'}), 401
    if request.method == 'GET':
        season = request.args.get('season', config.ACTIVE_SEASON)
        return jsonify(get_categories(season))
    elif request.method == 'POST':
        data = request.json
        name = data.get('name')
        season = data.get('season', config.ACTIVE_SEASON)
        if not name:
            return jsonify({'error': 'Название категории обязательно'}), 400
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        try:
            c.execute("INSERT OR IGNORE INTO categories (name, season) VALUES (?, ?)", (name, season))
            conn.commit()
            conn.close()
            return jsonify({'status': 'ok'})
        except Exception as e:
            conn.close()
            return jsonify({'error': str(e)}), 400
    elif request.method == 'DELETE':
        data = request.json
        name = data.get('name')
        if not name:
            return jsonify({'error': 'Название категории обязательно'}), 400
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        c.execute("DELETE FROM categories WHERE name=? AND season=?", (name, config.ACTIVE_SEASON))
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})

@app.route('/api/seasons', methods=['GET', 'POST', 'DELETE'])
def api_seasons():
    if 'user' not in session or session.get('role') != 'super_admin':
        return jsonify({'error': 'Unauthorized'}), 401
    if request.method == 'GET':
        return jsonify(get_seasons())
    elif request.method == 'POST':
        data = request.json
        name = data.get('name')
        if not name:
            return jsonify({'error': 'Название сезона обязательно'}), 400
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        try:
            c.execute("INSERT OR IGNORE INTO seasons (name) VALUES (?)", (name,))
            conn.commit()
            conn.close()
            return jsonify({'status': 'ok'})
        except Exception as e:
            conn.close()
            return jsonify({'error': str(e)}), 400
    elif request.method == 'DELETE':
        data = request.json
        name = data.get('name')
        if not name:
            return jsonify({'error': 'Название сезона обязательно'}), 400
        if name == config.ACTIVE_SEASON:
            return jsonify({'error': 'Нельзя удалить активный сезон'}), 400
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        c.execute("DELETE FROM clubs WHERE season=?", (name,))
        c.execute("DELETE FROM tours WHERE season=?", (name,))
        c.execute("DELETE FROM categories WHERE season=?", (name,))
        c.execute("DELETE FROM uploads WHERE season=?", (name,))
        c.execute("DELETE FROM seasons WHERE name=?", (name,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})

@app.route('/api/settings', methods=['GET', 'POST'])
def api_settings():
    if 'user' not in session or session.get('role') != 'super_admin':
        return jsonify({'error': 'Unauthorized'}), 401
    if request.method == 'GET':
        top_tours = get_top_tours()
        return jsonify({'top_tours': top_tours})
    elif request.method == 'POST':
        data = request.json
        top_tours = data.get('top_tours')
        if top_tours is None:
            return jsonify({'error': 'top_tours required'}), 400
        try:
            value = int(top_tours)
            if value < 1:
                return jsonify({'error': 'top_tours must be at least 1'}), 400
            set_top_tours(value)
            return jsonify({'status': 'ok', 'top_tours': value})
        except ValueError:
            return jsonify({'error': 'top_tours must be a number'}), 400

@app.route('/api/users', methods=['GET', 'POST', 'DELETE'])
def api_users():
    if 'user' not in session or session.get('role') != 'super_admin':
        return jsonify({'error': 'Unauthorized'}), 401
    if request.method == 'GET':
        return jsonify(get_users())
    elif request.method == 'POST':
        data = request.json
        username = data.get('username')
        password = data.get('password')
        club = data.get('club')
        if not username or not password or not club:
            return jsonify({'error': 'Логин, пароль и клуб обязательны'}), 400
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        try:
            c.execute("INSERT INTO users (username, password_hash, password_plain, role, club) VALUES (?, ?, ?, ?, ?)",
                      (username, password_hash, password, 'club_admin', club))
            conn.commit()
            conn.close()
            return jsonify({'status': 'ok', 'password': password})
        except Exception as e:
            conn.close()
            return jsonify({'error': str(e)}), 400
    elif request.method == 'DELETE':
        data = request.json
        username = data.get('username')
        if not username or username == 'Admin':
            return jsonify({'error': 'Нельзя удалить суперадмина'}), 400
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        c.execute("DELETE FROM users WHERE username=?", (username,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})

@app.route('/api/delete_upload/<int:upload_id>', methods=['DELETE'])
def delete_upload(upload_id):
    if 'user' not in session or session.get('role') != 'super_admin':
        return jsonify({'error': 'Unauthorized'}), 401
    
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    
    file_record = c.execute("SELECT filename FROM uploads WHERE id=?", (upload_id,)).fetchone()
    if not file_record:
        conn.close()
        return jsonify({'error': 'Upload not found'}), 404
    
    filename = file_record[0]
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    c.execute("DELETE FROM uploads WHERE id=?", (upload_id,))
    conn.commit()
    conn.close()
    
    if os.path.exists(filepath):
        try:
            os.remove(filepath)
        except:
            pass
    
    return jsonify({'status': 'ok'})

@app.route('/api/clear_all_uploads', methods=['DELETE'])
def clear_all_uploads():
    if 'user' not in session or session.get('role') != 'super_admin':
        return jsonify({'error': 'Unauthorized'}), 401
    
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    
    files = c.execute("SELECT filename FROM uploads").fetchall()
    
    deleted_files = 0
    for (filename,) in files:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(filepath):
            try:
                os.remove(filepath)
                deleted_files += 1
            except:
                pass
    
    c.execute("DELETE FROM uploads")
    conn.commit()
    conn.close()
    
    return jsonify({
        'status': 'ok',
        'deleted_records': len(files),
        'deleted_files': deleted_files
    })

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)